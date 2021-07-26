# _*_encoding:'utf-8' _*_
from openpyxl import Workbook, load_workbook

save_path = r"D:\Python" + r"\test.xlsx"
print("> create a new work book:")

wb = Workbook()#创建工作簿
print("> select sheet:")

wb.create_sheet("sheet1")#创建工作表
wb.create_sheet("sheet2")#创建工作表

ws = wb["sheet1"]#根据名字选择一个工作表
print(ws.title)

ws = wb.worksheets[0]#根据索引选择工作表
print(ws.title)
ws = wb.worksheets[1]#根据索引选择工作表
print(ws.title)

ws = wb.active#选择激活的工作表
print(ws.title)
print("> write data:")

#写入数据
for row in range(1, 5):
    for column in range(1, 5):
        ws.cell(row, column).value = str(row) + ',' + str(column)
        ws.cell(row + 5, column, str(row) + ',' + str(column))

#通过字符设置数据
ws["A1"] = 123
ws["A2"] = "test"
ws["B1"] = 1 + 2
ws["A1"].value = 123
ws["A2"].value = "test"
ws["B1"].value = 1 + 2

#通过字符获取数据
print(ws["A1"].value)
print(ws["A2"].value)
print(ws["B1"].value)

#通过切片获取范围
cell_range = ws["A1":"C3"]
print(cell_range)
for row in cell_range:
    for cell in row:
        print(cell.value)

print("> save work book:")
wb.save(save_path)
wb.close()

print("> load a existed work book:")
#加载工作簿
wb = load_workbook(save_path)
print("> print work book info:")

ws = wb.active#选择激活的工作表
print(ws.title)

print("roow:",ws.max_row, "column:",ws.max_column)

for row in range(1, ws.max_row + 1):
    for column in range(1, ws.max_column + 1):
        print(ws.cell(row, column).value, end=' ')
    print('')

#查看所有工作表
print(wb.sheetnames)

# wb.save("")#保存
# wb.close()#关闭

# ws["A1"],ws["A2"],ws["B1"]#获取单元格
# ws.cell(row,column)#获取单元格
# ws.cell(row,column,value)#设置单元格
# ws.cell(row,column).value#设置单元格
# ws["A1":"c3"]#切片获取一个范围内的单元格